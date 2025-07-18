import pandas as pd
from datetime import datetime
import getpass
import os
import platform
import pandas as pd
import numpy as np
import streamlit as st
import smtplib
import sys
import zipfile
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from dateutil import parser
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email.mime.image import MIMEImage
from email.utils import COMMASPACE, formatdate
from email import encoders
MailUSER = getpass.getuser()
MailHOST = platform.uname()[1]
MailFROM = 'fts_tagging@rn000116071.uhc.com'
MailSERVER = "mailo2.uhc.com"
def SendMail(email,MailSUBJECT, MailTEXT="", MailFILES=[]):
    MailTO = [email] 
    if type(MailTO) != list:
        MailTO = [MailTO]
    if type(MailFILES) != list:
        MailFILES = [MailFILES]

    msg = MIMEMultipart('related')
    msg['From'] = MailFROM
    msg['To'] = COMMASPACE.join(MailTO)
    msg['Date'] = formatdate(localtime=True)
    msg['Subject'] = MailSUBJECT

    # Create the body with HTML
    msgAlternative = MIMEMultipart('alternative')
    msg.attach(MIMEText(MailTEXT, "html"))
    msgText = MIMEText(MailTEXT, 'plain')
    msgAlternative.attach(msgText)
    # Attach image with Content-ID
    with open("UHC.png", 'rb') as img_file:
        img = MIMEImage(img_file.read())
        img.add_header('Content-ID', '<UHC.png>')
        img.add_header('Content-Disposition', 'inline', filename="UHC.png")
        msg.attach(img)

    # Attach files
    for file in MailFILES:
        part = MIMEBase('application', "octet-stream")
        part.set_payload(open(file, "rb").read())
        encoders.encode_base64(part)
        part.add_header('Content-Disposition', 'attachment; filename="%s"' % file)
        msg.attach(part)

    smtp = smtplib.SMTP(MailSERVER)
    smtp.sendmail(MailFROM, MailTO, msg.as_string())
    smtp.close()

    print("Email sent successfully!")

def main(df_source,df_target,email):
    print('source df:',df_source.shape)
    print('target df:',df_target.shape)

    # print("Source Columns:",df_source.columns)
    # print("Target Columns:",df_target.columns)

    df_target.columns = [col.replace('COS_', '') for col in df_target.columns]
    df_target.columns = [col.replace('_ID', '') for col in df_target.columns]
    df_target.rename(columns={'INCUR_DT':'GL_IDB_DT'}, inplace=True)
    df_target.rename(columns={'FINC_PRDCT_CD': 'FNC_PRDCT'}, inplace=True)
    df_target.rename(columns={'SRC_CUST_CONTR': 'GRP_POL_NBR'}, inplace=True)
    df_target.rename(columns={'PS9_GL_LE_CD':'PS9_GL_LEG_ENTY_CD'}, inplace=True)
    df_target.columns = [col.replace('CPTN_', 'CAP_') for col in df_target.columns]
    df_target.columns = [col.replace('CLM_', '') for col in df_target.columns]
    df_target.rename(columns={'CAP_DED_TYP_CD':'CAP_DED_TYP'}, inplace=True)
    df_target.rename(columns={'CLM_SRVC_TYP_CD':'SRVC_TYP_CD'}, inplace=True)
    df_source.columns = [col.replace('_A_', '_') for col in df_source.columns]

    df_target['LOOKUP_CODE_MASTER'] = None
    df_target['LOOKUP_CODE_DEPT'] = None
    df_target['LOOKUP_CODE_ACCOUNT'] = None
    df_target['LOOKUP_CODE_CUST'] = None
    if "GL_IDB_DT" in df_target.columns:
        df_target['GL_IDB_DT'] = pd.to_datetime(df_target['GL_IDB_DT']).dt.date

    common_cols = df_target.columns.intersection(df_source.columns)
    # Keep only common columns
    df_source = df_source[common_cols]
    df_target = df_target[common_cols]

    print('source df:',df_source.shape)
    print('target df:',df_target.shape)

    df_target = df_target[~df_target['DIV'].isin(['MSP', 'MEC', 'MTK'])]
    df_source = df_source[~df_source['DIV'].isin(['MSP', 'MEC', 'MTK'])]

    df_target.reset_index(drop=True, inplace=True)
    df_source.reset_index(drop=True, inplace=True)
    print("Shape after removing DIVs not in scope")
    print('source df:',df_source.shape)
    print('target df:',df_target.shape)

    ## here
    detailed_comparison = df_source.compare(df_target, keep_shape=True, keep_equal=True)
    detailed_comparison.fillna('DDD', inplace=True)
    compared = detailed_comparison[detailed_comparison.notna().any(axis=1)].astype(str)

    df = compared.copy()
    df.columns = df.columns.set_levels(['FTS_PreCheckGeneratedOutput' if x == 'self' else 'FTS_Tagged_Source' if x == 'other' else x for x in df.columns.levels[1]],level=1)
    df = df.drop(columns=[('LOOKUP_CODE_CUST', 'FTS_Tagged_Source'), ('LOOKUP_CODE_ACCOUNT', 'FTS_Tagged_Source'),('LOOKUP_CODE_DEPT', 'FTS_Tagged_Source'),('LOOKUP_CODE_MASTER', 'FTS_Tagged_Source')])

    # Compare and insert comparison columns next to each valid pair
    new_columns = []
    for col in df.columns.levels[0]:
        if ('FTS_PreCheckGeneratedOutput' in df[col]) and ('FTS_Tagged_Source' in df[col]):
            new_columns.append((col, 'FTS_PreCheckGeneratedOutput'))
            new_columns.append((col, 'FTS_Tagged_Source'))
            comparison_col = (col, 'Comparison')
            df[comparison_col] = df[(col, 'FTS_PreCheckGeneratedOutput')] == df[(col, 'FTS_Tagged_Source')]
            new_columns.append(comparison_col)
        else:
            new_columns.extend([(col, level) for level in df[col].columns])

    # Reorder columns
    df = df[new_columns]

    # Function to remove trailing alphabet from a string
    def remove_trailing_alpha(s):
        if isinstance(s, str) and s[-1].isalpha():
            return s[:-1]
        return s

    # Compare after cleaning 'FTS_PreCheckGeneratedOutput'
    df[('PS9_GL_ACCT_NBR', 'Comparison')] = df[('PS9_GL_ACCT_NBR', 'FTS_PreCheckGeneratedOutput')].apply(remove_trailing_alpha) == df[('PS9_GL_ACCT_NBR', 'FTS_Tagged_Source')]
    # Identify columns where the top-level does NOT start with 'PS9'
    mask = ~df.columns.get_level_values(0).str.startswith('PS9')
    # Apply replacement only to those columns
    df.loc[:, mask] = df.loc[:, mask].replace('DDD', '')
    fts_validate = 'ComparisonSheet_'+str(datetime.today().strftime('%Y%m%d_%H%M%S'))+'.xlsx'
    df.to_excel('tagged/'+fts_validate)
    # Ensure output directory exists
    output_path = 'tagged/'+fts_validate
    # Save to Excel
    df.to_excel(output_path)
    # Load workbook and worksheet
    wb = load_workbook(output_path)
    ws = wb.active
    # Define red fill
    red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
    # Highlight cells with value False
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is False:
                cell.fill = red_fill
    # Save workbook
    wb.save(output_path)
    print("Comparison Sheet created!")
    master_count = len(df[(df[('PS9_GL_LEG_ENTY_CD', 'Comparison')] == True) & (df[('PS9_GL_LOC_NBR', 'Comparison')] == True) & (df[('PS9_GL_PRDCT_CD', 'Comparison')] == True) & (df[('PS9_GL_SEG_CD', 'Comparison')] == True)])
    account_count = len(df[(df[('PS9_GL_ACCT_NBR', 'Comparison')] == True)])
    department_count = len(df[(df[('PS9_GL_DEPT_CD', 'Comparison')] == True)])
    customer_count = len(df[(df[('PS9_GL_CUST_NBR', 'Comparison')] == True)])

    mail_body = f'''
                <!DOCTYPE html>
                <html>
                <head>
                        <title>Comparison Summary</title>
                        <style>
                                table {{
                                        width: 80%;
                                        border-collapse: collapse;
                                        font-size: 14px;
                                }}
                                table, th, td {{
                                        border: 1px solid black;
                                }}
                                th, td {{
                                        padding: 10px;
                                        text-align: left;
                                }}
                                th {{
                                        background-color: #f2f2f2;
                                }}
                                p {{
                                        font-size: 14px;
                                        margin: 0;
                                }}
                                .small-table {{
                                        width: 50%;
                                        font-size: 12px;
                                        margin-top: 20px;
                                }}
                                .small-table th, .small-table td {{
                                        padding: 5px;
                                }}
                        </style>
                </head>
                <body>
                        <table style="border: none; margin-bottom: 20px; border-spacing: 0; border-collapse: collapse;" cellpadding="0" cellspacing="0">
                                <tr>
                                        <td style="border: none; vertical-align: middle; padding: 0; font-size: 0;">
                                                <img src="cid:UHC.png" alt="Logo" width="35" height="50" style="vertical-align: middle;">
                                        </td>
                                        <td style="border: none; vertical-align: middle; padding: 0; margin-left: 0;">
                                                <span style="font-size: 20px; font-weight: bold; margin-left: -5px; display: inline-block;">Comparison Summary</span>
                                        </td>
                                </tr>
                        </table>
                        <table>
                                <tr>
                                        <th>Lookup info</th>
                                        <th>Matched</th>
                                        <th>Not Matched</th>
                                </tr>
                                <tr>
                                        <td>Master Lookup</td>
                                        <td>{master_count}</td>
                                        <td>{len(df)-master_count}</td>
                                </tr>
                                <tr>
                                        <td>Account Lookup</td>
                                        <td>{account_count}</td>
                                        <td>{len(df)-account_count}</td>
                                </tr>
                                <tr>
                                        <td>Department Lookup</td>
                                        <td>{department_count}</td>
                                        <td>{len(df)-department_count}</td>
                                </tr>
                                <tr>
                                        <td>Customer Lookup</td>
                                        <td>{customer_count}</td>
                                        <td>{len(df)-customer_count}</td>
                                </tr>
                        </table>

                        <p>The complete results are attached as an Excel file named {fts_validate}.</p>
                        <p>Thank you</p>
                </body>
                </html>
                '''
    with zipfile.ZipFile('tagged/'+fts_validate+'.zip', 'w', zipfile.ZIP_DEFLATED) as zipf:
        zipf.write('tagged/'+fts_validate)
    SendMail(email,'FTS GL Tagging Comparison Report', mail_body, ['tagged/'+fts_validate+'.zip'])
    # SendMail(email,'FTS GL Tagging Comparison Report', mail_body, ['tagged/'+fts_validate])
    return df

if __name__ == "__main__":
    main(df_source,df_target,email)
